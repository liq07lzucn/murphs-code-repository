# coding: utf-8
from ROOT import TFile, TH1F, TCanvas
import os
from openpyxl import Workbook
import re

def GetRootFiles(path=os.getcwd(),gammaKey='Co60',neutronKey='neutron'):
  """ GetRootFiles

    Gets a list of root files, splitting up into gamma and neutrons
    according to gammaKey and neutronKey
  """
  gammaFiles = list()
  neutronFiles = list()
  for filename in os.listdir(path):
    [filename,ext] = os.path.splitext(filename)
    if ext == '.root':
      particleName = filename.split('_')[0].split('run')[0]
      filename = os.path.join(path,filename)+ext
      if particleName == gammaKey:
        gammaFiles.append(filename)
      elif particleName == neutronKey:
        neutronFiles.append(filename)
      else:
        print 'ROOT file '+filename+' is not reconized'
 return [gammaFiles,neutronFiles]

def PrintFiles(filelist,wb=Workbook(),histKey='eDepHist'):
  """ PlotFiles
  Prints the files files contained in filelist
  Keywords:
  fileName - name of the figure to save
  histKey - Key of the histogram in the root file
  """
  longName = ['Energy','Frequency','Error']
  units =['MeV','Particles per Event','Particles per Event']
  row = 0
  col = 0
  with open(fileName,'w') as outfile:
    for fname in filelist:
      f = TFile(fname,'r')
      hist = f.Get(histKey)
      # Writing the header

def GetThickness(filename):
  """ 
  GetThickness
  
  Parses a filename for a given thickness, returns the thickness in mm
  """
  value = filename.split('_')[-1]
  value = value.split('.')[0]
  # Uses a regular expression to seperate
  pattern = re.compile('(\d+)(\w+)')
  tokens = pattern.split(value)
  if tokens[2] == 'm':
    return float(tokens[1])*1000
  elif tokens[2] == 'cm':
    return float(tokens[1])*10
  elif token[2] == 'mm':
    return float(tokens[1])
  elif tokens[2] == 'um':
    return float(tokens[1])*0.001
  else:
    print tokens[2]," is not a reconized prefix"
    raise Exception()

def PrintEDepSummary(gFiles,nFiles,wb=Workbook(),tParse=GetThickness,
  histKey='eDepHist'):
  """ PrintEDepSummary
  Prints the energy deposition summary
  """
  ws = wb.create_sheet()
  ws.title = 'EDepSummary'
  # Extrating the average values
  nDict = dict()
  gDict = dict()
  for fname in gFiles:
    f = TFile(fname,'r')
    hist = f.Get(histKey)
    gDict[str(GetThickness(fname))] = (hist.GetMean(),hist.GetMeanError())

def main():
  print "Getting Files"
  [g,n] = GetRootFiles()
  print "Starting Data Analysis"
  wb = Workbook()
  PrintEDepSummary(g,n,wb)
  #print "Print Data"
  #PrintFiles(g,'GammaData.dat')
  wb.save('EnergyDepAnalysis.xlsx')

if __name__ == "__main__":
  main()
